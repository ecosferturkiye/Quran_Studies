import openpyxl
import json
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

# Source file
EXCEL_PATH = r'C:\Kitaplar\Modernways\Quran\Quran_Master_17052022.xlsx'
OUTPUT_DIR = r'C:\Users\aydin\OneDrive\Masaüstü\ClaudeCode\Test\next-linear-quran\src\data\quran-master'

# Create output directory
os.makedirs(OUTPUT_DIR, exist_ok=True)

print(f"Loading Excel file: {EXCEL_PATH}")
wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
ws = wb.active

# Get headers
headers = []
for col in range(1, ws.max_column + 1):
    headers.append(ws.cell(row=1, column=col).value)

print(f"Columns: {len(headers)}")
print(f"Total rows: {ws.max_row}")

# Process data by surah
surahs_data = {}
word_timing_data = {}

print("\nProcessing rows...")
row_count = 0

for row in ws.iter_rows(min_row=2, values_only=True):
    row_count += 1
    if row_count % 10000 == 0:
        print(f"  Processed {row_count} rows...")

    surah = row[0]  # surah
    ayah = row[1]   # ayah
    word_rank = row[2]  # word rank

    if surah is None:
        continue

    surah = int(surah)
    ayah = int(ayah) if ayah else 0
    word_rank = int(word_rank) if word_rank else 0

    # Initialize surah data
    if surah not in surahs_data:
        surahs_data[surah] = {
            "surahId": surah,
            "surahArabic": row[5] if row[5] else "",
            "surahLatin": row[6] if row[6] else "",
            "verses": {}
        }
        word_timing_data[surah] = {}

    # Initialize verse data
    verse_key = f"{surah}:{ayah}"
    if ayah not in surahs_data[surah]["verses"]:
        surahs_data[surah]["verses"][ayah] = {
            "verseNumber": ayah,
            "verseKey": verse_key,
            "wordCount": row[7] if row[7] else 0,
            "words": []
        }
        word_timing_data[surah][ayah] = []

    # Word data
    word_data = {
        "wordRank": word_rank,
        "arabic": row[26] if row[26] else "",  # arabic column (27 - 1 for 0-index)
        "tanzilClean": row[9] if row[9] else "",  # tanzil_clean
        "root": row[13] if row[13] else "",  # root
        "rootArabic": row[14] if row[14] else "",  # root_arabic
        "startTime": int(row[22]) if row[22] else 0,  # start_time
        "endTime": int(row[23]) if row[23] else 0,  # end_time
        "level1": row[16] if row[16] else None,  # level 1
        "level2": row[17] if row[17] else None,  # level 2
        "level3": row[18] if row[18] else None,  # level 3
        "level4": row[19] if row[19] else None,  # level 4
        "translations": {
            "en": row[27] if row[27] else "",  # translate_english
            "tr": row[32] if row[32] else "",  # translate_turkish
            "ur": row[28] if row[28] else "",  # translate_urdu
            "hi": row[29] if row[29] else "",  # translate_hindi
            "id": row[30] if row[30] else "",  # translate_indonesian
            "bn": row[31] if row[31] else "",  # translate_bangla
            "ru": row[33] if row[33] else "",  # translate_russian
        }
    }

    surahs_data[surah]["verses"][ayah]["words"].append(word_data)

    # Simplified timing data
    word_timing_data[surah][ayah].append({
        "w": word_rank,
        "s": word_data["startTime"],
        "e": word_data["endTime"]
    })

print(f"\nTotal rows processed: {row_count}")
print(f"Total surahs: {len(surahs_data)}")

# Save timing data (compact format for quick loading)
print("\nSaving word timing data...")
timing_output = {}
for surah_id, verses in word_timing_data.items():
    timing_output[surah_id] = {}
    for ayah, words in verses.items():
        timing_output[surah_id][ayah] = words

with open(os.path.join(OUTPUT_DIR, 'word-timing.json'), 'w', encoding='utf-8') as f:
    json.dump(timing_output, f, ensure_ascii=False)
print(f"  Saved: word-timing.json")

# Save full word data per surah (for magnifier feature)
print("\nSaving surah word data...")
for surah_id, surah_data in surahs_data.items():
    # Convert verses dict to list
    verses_list = []
    for ayah_num in sorted(surah_data["verses"].keys()):
        verses_list.append(surah_data["verses"][ayah_num])

    output_data = {
        "surahId": surah_data["surahId"],
        "surahArabic": surah_data["surahArabic"],
        "surahLatin": surah_data["surahLatin"],
        "verses": verses_list
    }

    filename = f'surah-{surah_id:03d}.json'
    with open(os.path.join(OUTPUT_DIR, filename), 'w', encoding='utf-8') as f:
        json.dump(output_data, f, ensure_ascii=False, indent=2)

print(f"  Saved {len(surahs_data)} surah files")

# Create index file
print("\nCreating index file...")
index_data = {
    "totalSurahs": len(surahs_data),
    "surahs": []
}

for surah_id in sorted(surahs_data.keys()):
    surah = surahs_data[surah_id]
    verse_count = len(surah["verses"])
    word_count = sum(len(v["words"]) for v in surah["verses"].values())

    index_data["surahs"].append({
        "id": surah_id,
        "nameArabic": surah["surahArabic"],
        "nameLatin": surah["surahLatin"],
        "verseCount": verse_count,
        "wordCount": word_count
    })

with open(os.path.join(OUTPUT_DIR, 'index.json'), 'w', encoding='utf-8') as f:
    json.dump(index_data, f, ensure_ascii=False, indent=2)
print("  Saved: index.json")

print("\nConversion complete!")
print(f"Output directory: {OUTPUT_DIR}")
